import copy
import csv
import io
import json
import re
from pathlib import Path

from django.conf import settings as django_settings
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.utils.translation import gettext as _
from django.contrib.auth.tokens import default_token_generator
from django.db.models import Q
from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_encode
from django.core.paginator import Paginator
from django.http import FileResponse, HttpResponse, StreamingHttpResponse
from django.shortcuts import get_object_or_404, redirect, render

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from kobo import api_client, cache_helpers
from kobo.models import KoboConfig, ConfiguredForm, DashboardGroup, DashboardConfig
from form_modules import get_module

PAGE_SIZE = 25
MODULES_DIR = Path(__file__).resolve().parent.parent / 'form_modules'


def _page_range(page_obj):
    """Return a compact list of page numbers and '...' separators for pagination."""
    total = page_obj.paginator.num_pages
    current = page_obj.number
    pages = set()
    pages.update([1, total])
    pages.update(range(max(1, current - 2), min(total + 1, current + 3)))
    result = []
    prev = None
    for p in sorted(pages):
        if prev is not None and p - prev > 1:
            result.append('...')
        result.append(p)
        prev = p
    return result

def _is_power_user(user):
    if not user.is_authenticated:
        return False
    if user.email in django_settings.POWER_USER_EMAILS:
        return True
    try:
        return user.profile.is_power_user
    except Exception:
        return False


def _is_group_admin(user):
    return user.is_authenticated and DashboardGroup.objects.filter(admins=user).exists()


def _user_accessible_forms(user):
    if _is_power_user(user):
        return ConfiguredForm.objects.all()
    return ConfiguredForm.objects.filter(
        Q(groups__members=user) | Q(groups__admins=user)
    ).distinct()


def _user_can_access_form(user, uid):
    if _is_power_user(user):
        return True
    return ConfiguredForm.objects.filter(uid=uid).filter(
        Q(groups__members=user) | Q(groups__admins=user)
    ).exists()


def _user_admin_forms(user):
    """Forms this user is group admin for (module upload/download)."""
    return ConfiguredForm.objects.filter(groups__admins=user).distinct()


def _can_manage_user(acting_user, target_pk):
    """True if acting_user may deactivate/delete/reset the target user."""
    if _is_power_user(acting_user):
        return True
    return DashboardGroup.objects.filter(
        admins=acting_user, members__pk=target_pk
    ).exists()


def _config():
    return KoboConfig.get()


def _get_form(uid):
    """Return ConfiguredForm for uid, or None."""
    try:
        return ConfiguredForm.objects.get(uid=uid)
    except ConfiguredForm.DoesNotExist:
        return None


def _load(uid):
    """Return (schema, submissions, structure, module) from cache."""
    config = _config()
    form = _get_form(uid)
    ttl = form.cache_ttl_seconds if form else 300

    schema = cache_helpers.get_cached(
        cache_helpers.schema_key(uid),
        lambda: api_client.get_schema(uid, config),
        ttl=ttl,
    )
    submissions = cache_helpers.get_cached(
        cache_helpers.submissions_key(uid),
        lambda: api_client.get_submissions(uid, config),
        ttl=ttl,
    )
    module = get_module(uid)
    if module is not None:
        structure = cache_helpers.get_cached(
            f'kobo_structure_{uid}',
            lambda: module.parse_structure(schema),
            ttl=ttl,
        )
    else:
        structure = {}
    return schema, submissions, structure, module


# ── Form list ──────────────────────────────────────────────────────────────────

@login_required
def form_list(request):
    forms = _user_accessible_forms(request.user)
    if not forms.exists():
        if _is_power_user(request.user):
            return redirect('/dashboard/settings/')
        return render(request, 'dashboard/no_form.html', {})

    is_power_user = _is_power_user(request.user)
    admin_uids = set(
        _user_admin_forms(request.user).values_list('uid', flat=True)
    ) if not is_power_user else None

    form_cards = []
    for f in forms:
        module = get_module(f.uid)
        cached_subs = cache_helpers.get_if_cached(cache_helpers.submissions_key(f.uid))
        if module:
            from pathlib import Path as _Path
            module_name = module.form_label or _Path(module._source_file).stem
        else:
            module_name = None
        can_edit = is_power_user or f.uid in admin_uids
        form_cards.append({
            'uid': f.uid,
            'name': f.name,
            'module_label': module.form_label if module else None,
            'module_name': module_name,
            'sub_count': len(cached_subs) if cached_subs is not None else None,
            'dash_configs': list(f.dashboard_configs.values('id', 'name')),
            'can_edit': can_edit,
        })

    return render(request, 'dashboard/form_list.html', {
        'form_cards': form_cards,
        'is_power_user': is_power_user,
    })


# ── Settings ───────────────────────────────────────────────────────────────────

@login_required
def settings_view(request):
    if not _is_power_user(request.user):
        return redirect('/dashboard/')

    config = _config()
    assets = []
    error = None
    success = None
    show_add_form = False

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'save_server':
            server_url = request.POST.get('server_url', '').strip().rstrip('/')
            api_token = request.POST.get('api_token', '').strip()
            if server_url:
                config.server_url = server_url
            if api_token:
                config.api_token = api_token
            config.save()
            cache_helpers.invalidate(cache_helpers.asset_list_key())
            success = _('Connexion enregistrée.')

        elif action == 'load_assets':
            server_url = request.POST.get('server_url', '').strip().rstrip('/')
            api_token = request.POST.get('api_token', '').strip()
            if server_url:
                config.server_url = server_url
            if api_token:
                config.api_token = api_token
            config.save()
            cache_helpers.invalidate(cache_helpers.asset_list_key())
            try:
                assets = api_client.list_assets(config)
                show_add_form = True
            except api_client.KoboAPIError as exc:
                error = str(exc)

        elif action == 'add_form':
            uid = request.POST.get('selected_form_uid', '').strip()
            name = request.POST.get('selected_form_name', '').strip()
            ttl = request.POST.get('cache_ttl_seconds', '300').strip()
            if uid and not ConfiguredForm.objects.filter(uid=uid).exists():
                ConfiguredForm.objects.create(
                    uid=uid,
                    name=name or uid,
                    cache_ttl_seconds=int(ttl) if ttl.isdigit() else 300,
                    order=ConfiguredForm.objects.count(),
                )
                success = _('Formulaire « %(name)s » ajouté.') % {'name': name}
            elif uid:
                error = _('Ce formulaire est déjà configuré.')

        elif action == 'remove_form':
            uid = request.POST.get('form_uid', '').strip()
            ConfiguredForm.objects.filter(uid=uid).delete()
            for key in [cache_helpers.schema_key(uid),
                        cache_helpers.submissions_key(uid),
                        f'kobo_structure_{uid}']:
                cache_helpers.invalidate(key)
            success = _('Formulaire supprimé.')

        elif action == 'update_ttl':
            uid = request.POST.get('form_uid', '').strip()
            ttl = request.POST.get('cache_ttl_seconds', '300').strip()
            ConfiguredForm.objects.filter(uid=uid).update(
                cache_ttl_seconds=int(ttl) if ttl.isdigit() else 300
            )
            success = _('Durée du cache mise à jour.')

        elif action == 'save_branding':
            config.brand_color = request.POST.get('brand_color', '').strip()
            config.org_name = request.POST.get('org_name', '').strip()
            if 'logo' in request.FILES:
                config.logo = request.FILES['logo']
            elif request.POST.get('remove_logo'):
                config.logo.delete(save=False)
                config.logo = None
            config.save()
            success = _('Apparence enregistrée.')

        elif action == 'create_group':
            gname = request.POST.get('group_name', '').strip()
            if gname:
                _group, created = DashboardGroup.objects.get_or_create(name=gname)
                success = (_('Groupe « %(name)s » créé.') % {'name': gname}) if created else _('Ce nom de groupe existe déjà.')

        elif action == 'delete_group':
            gid = request.POST.get('group_id', '').strip()
            DashboardGroup.objects.filter(pk=gid).delete()
            success = _('Groupe supprimé.')

    configured_forms = []
    for f in ConfiguredForm.objects.all():
        module = get_module(f.uid)
        configured_forms.append({
            'uid': f.uid,
            'name': f.name,
            'cache_ttl_seconds': f.cache_ttl_seconds,
            'module_label': module.form_label if module else None,
            'has_module_file': module is not None,
        })

    groups = DashboardGroup.objects.prefetch_related('forms', 'members', 'admins').all()

    return render(request, 'dashboard/settings.html', {
        'config': config,
        'configured_forms': configured_forms,
        'assets': assets,
        'show_add_form': show_add_form,
        'error': error,
        'success': success,
        'groups': groups,
    })


# ── Module download / upload ───────────────────────────────────────────────────

@login_required
def module_download(request, uid):
    if not _is_power_user(request.user) and uid not in _user_admin_forms(request.user).values_list('uid', flat=True):
        return redirect('/dashboard/')
    module = get_module(uid)
    if module is None:
        return HttpResponse('Aucun module pour ce formulaire.', status=404)
    path = Path(module._source_file)
    return FileResponse(open(path, 'rb'), as_attachment=True, filename=path.name)


@login_required
def module_upload(request, uid):
    can_upload = _is_power_user(request.user) or uid in _user_admin_forms(request.user).values_list('uid', flat=True)
    if not can_upload or request.method != 'POST':
        return redirect('/dashboard/')

    uploaded = request.FILES.get('module_file')
    if not uploaded:
        return redirect('/dashboard/settings/')

    stem = Path(uploaded.name).stem
    if not re.match(r'^[A-Za-z][A-Za-z0-9_]*$', stem):
        return render(request, 'dashboard/settings.html',
                      {'error': _('Nom de fichier invalide (doit être un identifiant Python valide).'),
                       'config': _config(), 'configured_forms': [], 'assets': [], 'show_add_form': False, 'success': None})

    dest = MODULES_DIR / f'{stem}.py'
    dest.write_bytes(uploaded.read())

    return render(request, 'dashboard/module_uploaded.html', {
        'filename': dest.name,
        'uid': uid,
    })


def _build_table_rows(filtered, f_result):
    """Return flat list of indicator rows for the data table."""
    rows = []
    for ps in filtered:
        for ind in ps['indicators']:
            if f_result and ind['result_key'] != f_result:
                continue
            rows.append({
                'country': ps['country_label'],
                'period': ps['period'],
                'reporter': ps['reporter'],
                'result': ind['result_label'],
                'code': ind['code'],
                'label': ind['label'],
                'total': ind['total'],
                'male': ind['age'].get('male_total', ''),
                'female': ind['age'].get('fem_total', ''),
                'disability': ind['disability'].get('with', ''),
                'pdi': ind['status'].get('pdi', ''),
            })
    return rows


# ── Generic form detail (fallback for forms with no module) ───────────────────

@login_required
def module_dashboard(request, uid):
    """Dashboard for indicator-monitoring form modules (modules with parse_submissions)."""
    if not _user_can_access_form(request.user, uid):
        return redirect('/dashboard/')
    error = None
    chart_data = {}
    parsed_all = []
    filtered = []
    total_beneficiaries = 0
    total_reports = 0
    countries_used = []
    results_used = []
    country_labels = {}
    result_labels = {}
    module = None

    f_country = request.GET.get('country', '')
    f_year = request.GET.get('year', '')
    f_quarter = request.GET.get('quarter', '')
    f_result = request.GET.get('result', '')

    try:
        schema, submissions, structure, module = _load(uid)
        if module is None or not hasattr(module, 'parse_submissions'):
            return form_detail(request, uid)

        country_labels = getattr(module, 'COUNTRY_LABELS', {})
        result_labels = getattr(module, 'RESULT_LABELS', {})
        indicator_labels = getattr(module, 'INDICATOR_LABELS', {})
        result_keys = getattr(module, 'RESULT_KEYS', [])
        country_colors = getattr(module, 'COUNTRY_COLORS', {})
        aggregate = getattr(module, 'aggregate', None)

        parsed_all = module.parse_submissions(submissions)

        # Collect filter options from data
        countries_used = sorted({ps['country'] for ps in parsed_all if ps['country']})
        results_used_set = set()
        for ps in parsed_all:
            for ind in ps['indicators']:
                results_used_set.add(ind['result_key'])
        results_used = [r for r in (result_keys or sorted(results_used_set))
                        if r in results_used_set]

        # Apply filters
        filtered = parsed_all
        if f_country:
            filtered = [ps for ps in filtered if ps['country'] == f_country]
        if f_year:
            filtered = [ps for ps in filtered if ps['year'] == f_year]
        if f_quarter:
            filtered = [ps for ps in filtered if ps['quarter'] == f_quarter]

        total_reports = len(filtered)

        agg = {'by_country': {}, 'by_indicator': {}, 'by_period': {}}
        if aggregate is not None:
            agg = aggregate(filtered)
        total_beneficiaries = sum(agg['by_country'].values())

        country_list = [c for c in country_labels if c in countries_used]
        if f_country:
            country_list = [f_country] if f_country in country_labels else []

        result_charts = []
        for rkey in result_keys:
            if f_result and rkey != f_result:
                continue
            ind_codes = []
            for ps in filtered:
                for ind in ps['indicators']:
                    if ind['result_key'] == rkey and ind['code'] not in ind_codes:
                        ind_codes.append(ind['code'])
            if not ind_codes:
                continue
            ind_labels = [indicator_labels.get(c, c) for c in ind_codes]
            datasets = []
            for country in country_list:
                data = [agg['by_indicator'].get(code, {}).get(country, 0) for code in ind_codes]
                datasets.append({
                    'label': country_labels.get(country, country),
                    'data': data,
                    'backgroundColor': country_colors.get(country, '#999'),
                })
            result_charts.append({
                'result_key': rkey,
                'result_label': result_labels.get(rkey, rkey),
                'indicator_labels': ind_labels,
                'indicator_codes': ind_codes,
                'datasets': datasets,
            })

        country_summary = {
            'labels': [country_labels.get(c, c) for c in country_list],
            'data': [agg['by_country'].get(c, 0) for c in country_list],
            'colors': [country_colors.get(c, '#999') for c in country_list],
        }
        period_labels = sorted({ps['period'] for ps in filtered if ps['period']})
        period_data = [agg['by_period'].get(p, 0) for p in period_labels]

        age_totals = {'male_0_5': 0, 'male_6_18': 0, 'male_19_49': 0, 'male_50p': 0,
                      'fem_0_5': 0, 'fem_6_18': 0, 'fem_19_49': 0, 'fem_50p': 0}
        disability_totals = {'with': 0, 'without': 0}
        status_totals = {'pdi': 0, 'host': 0, 'refugee': 0, 'returnees': 0,
                         'stateless': 0, 'other': 0}
        has_disagg = False
        for ps in filtered:
            for ind in ps['indicators']:
                if f_result and ind['result_key'] != f_result:
                    continue
                if ind['age']:
                    has_disagg = True
                    for k in age_totals:
                        age_totals[k] += ind['age'].get(k, 0)
                if ind['disability']:
                    for k in disability_totals:
                        disability_totals[k] += ind['disability'].get(k, 0)
                if ind['status']:
                    for k in status_totals:
                        status_totals[k] += ind['status'].get(k, 0)

        disagg_chart = None
        if has_disagg:
            age_labels = getattr(module, 'AGE_DISAGG_LABELS',
                                 ['0–5 M', '6–18 M', '19–49 M', '50+ M',
                                  '0–5 F', '6–18 F', '19–49 F', '50+ F'])
            age_data = [age_totals[k] for k in age_totals]
            age_colors = getattr(module, 'AGE_DISAGG_COLORS',
                                 (['#156082'] * 4) + (['#c00000'] * 4))
            status_labels = getattr(module, 'STATUS_LABELS',
                                    ['PDI', 'Hôte', 'Réfugié', 'Rapatrié', 'Migrant', 'Autre'])
            status_data = [status_totals[k] for k in status_totals]
            disagg_chart = {
                'age': {'labels': age_labels, 'data': age_data, 'colors': age_colors},
                'disability': {
                    'labels': ['Avec handicap', 'Sans handicap'],
                    'data': [disability_totals['with'], disability_totals['without']],
                    'colors': ['#e97132', '#196b24'],
                },
                'status': {'labels': status_labels, 'data': status_data},
            }

        chart_data = {
            'country_summary': country_summary,
            'period_trend': {'labels': period_labels, 'data': period_data},
            'result_charts': result_charts,
            'disagg': disagg_chart,
        }

    except api_client.KoboAPIError as exc:
        error = str(exc)

    template = getattr(module, 'dashboard_template', 'dashboard/module_dashboard.html')
    return render(request, template, {
        'uid': uid,
        'form_label': getattr(module, 'form_label', uid),
        'error': error,
        'total_beneficiaries': total_beneficiaries,
        'total_reports': total_reports,
        'countries_used': countries_used,
        'results_used': results_used,
        'periods': sorted({ps['period'] for ps in parsed_all if ps['period']}),
        'years': sorted({ps['year'] for ps in parsed_all if ps['year']}),
        'quarters': ['Q1', 'Q2', 'Q3', 'Q4'],
        'f_country': f_country,
        'f_year': f_year,
        'f_quarter': f_quarter,
        'f_result': f_result,
        'country_labels': country_labels,
        'result_labels': result_labels,
        'chart_data_json': json.dumps(chart_data),
        'table_rows': _build_table_rows(filtered, f_result),
    })


@login_required
def form_detail(request, uid):
    if not _user_can_access_form(request.user, uid):
        return redirect('/dashboard/')
    error = None
    tabs = []
    group_tree = []
    columns = []
    page_obj = None
    form_name = uid
    total_submissions = 0
    active_group = ''

    try:
        config = _config()
        schema = cache_helpers.get_cached(
            cache_helpers.schema_key(uid),
            lambda: api_client.get_schema(uid, config),
        )
        submissions = cache_helpers.get_cached(
            cache_helpers.submissions_key(uid),
            lambda: api_client.get_submissions(uid, config),
        )
        form_name = schema.get('name', uid)
        total_submissions = len(submissions)

        group_order, groups = api_client.parse_groups(schema)
        question_labels = api_client.get_question_labels(schema)

        tabs = [{'key': k, 'label': groups[k]['label']} for k in group_order]
        group_tree = api_client.parse_group_tree(schema, set(group_order))
        active_group = request.GET.get('group', group_order[0] if group_order else '')
        if active_group not in groups and group_order:
            active_group = group_order[0]

        if active_group and active_group in groups:
            group_info = groups[active_group]
            questions = group_info['questions']
            columns = [{'label': question_labels.get(q, q)} for q in questions]

            is_repeat = group_info.get('is_repeat', False)
            parent_rpt_key = group_info.get('parent_repeat_key')  # full_key of enclosing repeat
            full_key = group_info.get('full_key') or active_group

            rows = []
            if is_repeat:
                if parent_rpt_key:
                    # Nested repeat: iterate parent repeat items, then nested items
                    for sub in submissions:
                        for parent_item in sub.get(parent_rpt_key, []):
                            for item in parent_item.get(full_key, []):
                                rows.append({'id': sub.get('_id', ''),
                                             'values': [item.get(q, '') for q in questions]})
                else:
                    # Top-level repeat: iterate repeat items
                    for sub in submissions:
                        for item in sub.get(full_key, []):
                            rows.append({'id': sub.get('_id', ''),
                                         'values': [item.get(q, '') for q in questions]})
            elif parent_rpt_key:
                # Regular group inside a repeat: questions live in repeat items
                for sub in submissions:
                    for parent_item in sub.get(parent_rpt_key, []):
                        rows.append({'id': sub.get('_id', ''),
                                     'values': [parent_item.get(q, '') for q in questions]})
            else:
                # Regular top-level group: flat submission lookup
                rows = [
                    {'id': sub.get('_id', ''), 'values': [sub.get(q, '') for q in questions]}
                    for sub in submissions
                ]

            paginator = Paginator(rows, PAGE_SIZE)
            page_obj = paginator.get_page(request.GET.get('page'))

    except api_client.KoboAPIError as exc:
        error = str(exc)

    return render(request, 'dashboard/form_detail.html', {
        'uid': uid,
        'form_name': form_name,
        'total_submissions': total_submissions,
        'tabs': tabs,
        'group_tree': group_tree,
        'active_group': active_group,
        'columns': columns,
        'page_obj': page_obj,
        'page_range': _page_range(page_obj) if page_obj else [],
        'error': error,
    })


# ── Generic JSON dashboard renderer ───────────────────────────────────────────

_WIDGET_COLORS = [
    '#dc3545', '#0d6efd', '#198754', '#fd7e14',
    '#6f42c1', '#20c997', '#ffc107', '#0dcaf0',
]

_COL_CLASS = {1: 'col-12', 2: 'col-md-6', 3: 'col-md-4'}


def _apply_widget_filters(submissions, widget_filters):
    def _match(sub, f):
        val = str(sub.get(f.get('field', ''), '') or '')
        op, target = f.get('op', 'eq'), f.get('value', '')
        if op == 'eq':
            return val == target
        if op == 'neq':
            return val != target
        if op == 'empty':
            return not val
        if op == 'notempty':
            return bool(val)
        return True
    return [s for s in submissions if all(_match(s, f) for f in widget_filters)]


def _render_widget(widget, submissions, schema):
    wtype = widget.get('type', '')
    title = widget.get('title', '')

    widget_filters = widget.get('widget_filters', [])
    if widget_filters:
        submissions = _apply_widget_filters(submissions, widget_filters)

    if wtype == 'summary_stat':
        field = widget.get('field') or None
        agg = widget.get('aggregation', 'count')
        if agg == 'count' or not field:
            value = len(submissions)
        else:
            nums = []
            for s in submissions:
                try:
                    nums.append(float(s.get(field, '')))
                except (TypeError, ValueError):
                    pass
            if agg == 'sum':
                value = int(sum(nums))
            else:
                value = round(sum(nums) / len(nums), 1) if nums else 0
        data = {'value': value}

    elif wtype == 'pie_chart':
        field = widget.get('field', '')
        choice_labels = api_client.get_choice_labels(schema, field) if field else {}
        pie_meta = widget.get('pie_meta', {})
        counts = {}
        for sub in submissions:
            val = sub.get(field, '')
            if val:
                counts[val] = counts.get(val, 0) + 1
        sorted_items = sorted(counts.items(), key=lambda x: -x[1])
        labels = [
            pie_meta.get(v, {}).get('label') or choice_labels.get(v, v)
            for v, _ in sorted_items
        ]
        values = [c for _, c in sorted_items]
        colors = [
            pie_meta.get(v, {}).get('color') or _WIDGET_COLORS[i % len(_WIDGET_COLORS)]
            for i, (v, _) in enumerate(sorted_items)
        ]
        data = {'labels': labels, 'values': values, 'colors': colors}

    elif wtype == 'data_table':
        question_labels = api_client.get_question_labels(schema)
        if widget.get('pivot_mode'):
            row_field = widget.get('pivot_row_field', '')
            col_field = widget.get('pivot_col_field', '')
            cl_row = api_client.get_choice_labels(schema, row_field) if row_field else {}
            cl_col = api_client.get_choice_labels(schema, col_field) if col_field else {}
            row_order, row_seen = [], set()
            col_order, col_seen = [], set()
            for sub in submissions:
                rv = str(sub.get(row_field, '') or '')
                cv = str(sub.get(col_field, '') or '')
                if rv and rv not in row_seen:
                    row_seen.add(rv); row_order.append(rv)
                if cv and cv not in col_seen:
                    col_seen.add(cv); col_order.append(cv)
            counts = {rv: {cv: 0 for cv in col_order} for rv in row_order}
            row_totals = {rv: 0 for rv in row_order}
            col_totals = {cv: 0 for cv in col_order}
            for sub in submissions:
                rv = str(sub.get(row_field, '') or '')
                cv = str(sub.get(col_field, '') or '')
                if rv in row_seen and cv in col_seen:
                    counts[rv][cv] += 1
                    row_totals[rv] += 1
                    col_totals[cv] += 1
            data = {
                'pivot': True,
                'row_label': question_labels.get(row_field, row_field),
                'col_label': question_labels.get(col_field, col_field),
                'col_headers': [cl_col.get(cv, cv) for cv in col_order],
                'rows': [
                    {
                        'label': cl_row.get(rv, rv),
                        'cells': [counts[rv][cv] for cv in col_order],
                        'total': row_totals[rv],
                    }
                    for rv in row_order
                ],
                'col_totals': [col_totals[cv] for cv in col_order],
                'grand_total': sum(row_totals.values()),
            }
        else:
            fields = widget.get('fields', [])
            headers = [question_labels.get(f, f) for f in fields]
            rows_data = [[str(sub.get(f, '')) for f in fields] for sub in submissions[:200]]
            data = {'headers': headers, 'rows': rows_data}

    elif wtype == 'grouped_chart':
        field1 = widget.get('field', '')
        chart_style = widget.get('chart_style', 'bar')
        stacked = bool(widget.get('stacked', False))
        custom_labels = widget.get('custom_labels', {})
        cl1 = api_client.get_choice_labels(schema, field1) if field1 else {}
        qlabels = api_client.get_question_labels(schema) if schema else {}
        # Normalise series to list of dicts {field, color, label}
        raw_series = widget.get('series') or ([widget['field2']] if widget.get('field2') else [])
        series_configs = []
        for idx, entry in enumerate(raw_series):
            if isinstance(entry, dict):
                series_configs.append({
                    'field': entry.get('field', ''),
                    'color': entry.get('color') or _WIDGET_COLORS[idx % len(_WIDGET_COLORS)],
                    'label': entry.get('label', ''),
                })
            else:
                series_configs.append({
                    'field': entry,
                    'color': _WIDGET_COLORS[idx % len(_WIDGET_COLORS)],
                    'label': '',
                })
        series_fields = [s['field'] for s in series_configs if s['field']]
        multi_sf = len(series_fields) > 1
        color_idx = 0
        series_value_colors = widget.get('series_value_colors') or {}

        def _unique_vals(field):
            order, seen = [], set()
            for sub in submissions:
                v = str(sub.get(field, '') or '')
                if v and v not in seen:
                    seen.add(v); order.append(v)
            return order, seen

        if field1:
            # X axis = values of field1
            labels_order, seen_l = _unique_vals(field1)
            if series_fields:
                # Cross-tab: one dataset per unique value of each series field
                datasets = []
                for sc in series_configs:
                    sf = sc['field']
                    if not sf:
                        continue
                    cl2 = api_client.get_choice_labels(schema, sf)
                    series_order, seen_s = _unique_vals(sf)
                    counts = {s: {l: 0 for l in labels_order} for s in series_order}
                    for sub in submissions:
                        v1 = str(sub.get(field1, '') or '')
                        v2 = str(sub.get(sf, '') or '')
                        if v1 in seen_l and v2 in seen_s:
                            counts[v2][v1] += 1
                    path_svc = series_value_colors.get(sf, {})
                    series_base_color = sc['color']
                    for val_idx, sk in enumerate(series_order):
                        lbl = cl2.get(sk, sk)
                        if multi_sf:
                            field_lbl = sc['label'] or qlabels.get(sf, sf)
                            lbl = f"{field_lbl} — {lbl}"
                        if path_svc.get(sk):
                            color = path_svc[sk]
                        elif val_idx == 0:
                            color = series_base_color
                        else:
                            color = _WIDGET_COLORS[color_idx % len(_WIDGET_COLORS)]
                        color_idx += 1
                        datasets.append({
                            'label': lbl,
                            'data': [counts[sk][l] for l in labels_order],
                            'backgroundColor': color, 'borderColor': color, 'fill': False,
                        })
            else:
                # Simple distribution: count per field1 value, no series
                counts_s = {}
                for sub in submissions:
                    v = str(sub.get(field1, '') or '')
                    if v: counts_s[v] = counts_s.get(v, 0) + 1
                color = (series_configs[0]['color'] if series_configs else _WIDGET_COLORS[0])
                datasets = [{'label': qlabels.get(field1, field1),
                             'data': [counts_s.get(l, 0) for l in labels_order],
                             'backgroundColor': color, 'borderColor': color, 'fill': False}]
            x_labels = [custom_labels.get(l, cl1.get(l, l)) for l in labels_order]
        else:
            # No X field: flat absolute comparison — one dataset per series field,
            # X axis = all (field, value) pairs concatenated
            x_labels, datasets = [], []
            for sc in series_configs:
                sf = sc['field']
                if not sf:
                    continue
                cl2 = api_client.get_choice_labels(schema, sf)
                series_order, _ = _unique_vals(sf)
                counts_s = {}
                for sub in submissions:
                    v = str(sub.get(sf, '') or '')
                    if v: counts_s[v] = counts_s.get(v, 0) + 1
                start = len(x_labels)
                x_labels.extend(cl2.get(v, v) for v in series_order)
                data_arr = [None] * start + [counts_s.get(v, 0) for v in series_order]
                color = sc['color']
                color_idx += 1
                field_lbl = sc['label'] or qlabels.get(sf, sf)
                datasets.append({'label': field_lbl, 'data': data_arr,
                                 'backgroundColor': color, 'borderColor': color, 'fill': False})
            # Pad all datasets to same length
            for ds in datasets:
                ds['data'] += [None] * (len(x_labels) - len(ds['data']))
        data = {
            'chart_style': chart_style, 'stacked': stacked,
            'labels': x_labels, 'datasets': datasets,
        }

    else:
        data = {}

    return {
        'type': wtype,
        'title': title,
        'data': data,
        'data_json': json.dumps(data, ensure_ascii=False),
    }


def _render_generic_dashboard(request, uid, form, dash_config):
    error = None
    rows_rendered = []
    filter_bars = []
    total_submissions = 0
    filtered_count = 0
    config_json = dash_config.config or {}

    try:
        kobo_config = _config()
        ttl = form.cache_ttl_seconds
        schema = cache_helpers.get_cached(
            cache_helpers.schema_key(uid),
            lambda: api_client.get_schema(uid, kobo_config),
            ttl=ttl,
        )
        submissions = cache_helpers.get_cached(
            cache_helpers.submissions_key(uid),
            lambda: api_client.get_submissions(uid, kobo_config),
            ttl=ttl,
        )
        total_submissions = len(submissions)

        # Build filter bars and apply active filters
        filters_config = config_json.get('filters', [])
        active_filters = {}
        for f in filters_config:
            field = f.get('field', '')
            val = request.GET.get(field, '')
            if val:
                active_filters[field] = val

        filtered = [
            s for s in submissions
            if all(s.get(field) == val for field, val in active_filters.items())
        ]
        filtered_count = len(filtered)

        for f in filters_config:
            field = f.get('field', '')
            label = f.get('label') or field
            choice_labels = api_client.get_choice_labels(schema, field) if field else {}
            distinct_vals = sorted({s.get(field, '') for s in submissions if s.get(field, '')})
            filter_bars.append({
                'field': field,
                'label': label,
                'active': active_filters.get(field, ''),
                'options': [{'value': v, 'label': choice_labels.get(v, v)} for v in distinct_vals],
            })

        for i, row in enumerate(config_json.get('rows', [])):
            widgets_rendered = []
            for j, widget in enumerate(row.get('widgets', [])):
                rendered = _render_widget(widget, filtered, schema)
                rendered['canvas_id'] = f'chart_{i}_{j}'
                widgets_rendered.append(rendered)
            columns = max(row.get('columns', 1), min(len(widgets_rendered), 3))
            rows_rendered.append({
                'columns': columns,
                'col_class': _COL_CLASS.get(columns, 'col-12'),
                'widgets': widgets_rendered,
            })
    except api_client.KoboAPIError as exc:
        error = str(exc)

    can_edit = _is_power_user(request.user) or \
        uid in _user_admin_forms(request.user).values_list('uid', flat=True)

    return render(request, 'dashboard/generic_dashboard.html', {
        'uid': uid,
        'form_name': form.name,
        'dashboard_name': dash_config.name,
        'dashboard_pk': dash_config.pk,
        'rows': rows_rendered,
        'filter_bars': filter_bars,
        'total_submissions': total_submissions,
        'filtered_count': filtered_count,
        'has_filters': bool(filter_bars),
        'error': error,
        'can_edit': can_edit,
    })


# ── Coverage matrix ────────────────────────────────────────────────────────────

@login_required
def coverage(request, uid):
    if not _user_can_access_form(request.user, uid):
        return redirect('/dashboard/')

    # ?group= means the user is navigating raw-data tabs — go straight to form_detail.
    if request.GET.get('group'):
        return form_detail(request, uid)

    # Build list of all available dashboard options (JSON dashboards + Python module).
    # Skip selection when ?view=module is present (direct module link).
    if request.GET.get('view') != 'module':
        _module = get_module(uid)
        try:
            _cf = ConfiguredForm.objects.get(uid=uid)
            _json_dbs = list(_cf.dashboard_configs.order_by('pk'))
        except ConfiguredForm.DoesNotExist:
            _json_dbs = []
            _cf = None

        _options = []
        for _d in _json_dbs:
            _options.append({'type': 'json', 'pk': _d.pk, 'name': _d.name or str(_d.pk)})
        _module_inst = _module
        if _module_inst is not None:
            _label = getattr(_module_inst, 'form_label', '') or uid
            _mtype = 'indicator' if hasattr(_module_inst, 'parse_submissions') else 'coverage_matrix'
            _options.append({'type': _mtype, 'name': _label})

        if len(_options) > 1:
            return render(request, 'dashboard/dashboard_select.html',
                          {'options': _options, 'uid': uid, 'form': _cf})

        if len(_options) == 1 and _options[0]['type'] == 'json':
            return redirect('view_dashboard', uid=uid, pk=_options[0]['pk'])

    error = None
    structure = {}
    coverage_data = {}
    responsibles = []
    module = None

    f_country = request.GET.get('country', '')
    f_result = request.GET.get('result', '')
    f_activity = request.GET.get('activity', '')
    f_responsible = request.GET.get('responsible', '')

    try:
        schema, submissions, structure, module = _load(uid)
    except api_client.KoboAPIError as exc:
        error = str(exc)

    if module is None and not error:
        return form_detail(request, uid)

    # Route modules that have their own dashboard view
    if module is not None and hasattr(module, 'parse_submissions') and not error:
        return module_dashboard(request, uid)

    if not error:
        fp = module.FIELD_PATHS
        resp_by_country = {}
        for sub in submissions:
            act_code = sub.get(fp['activity_code'], '')
            country = sub.get(fp['country'], '')
            main = module.extract_main_activity(act_code)
            responsible = sub.get(fp['activity_responsible'], '').strip()
            if main and country:
                coverage_data[(main, country)] = coverage_data.get((main, country), 0) + 1
            if country and responsible:
                resp_by_country.setdefault(country, set()).add(responsible)

        if f_country:
            resp_set = resp_by_country.get(f_country, set())
        else:
            resp_set = {name for names in resp_by_country.values() for name in names}
        responsibles = sorted(resp_set)

    applicable = structure.get('applicable', set())
    results = structure.get('results', [])
    if f_result:
        results = [r for r in results if r['code'] == f_result]
    if f_country:
        results = [
            {**r, 'activities': [
                a for a in r['activities']
                if (a['code'], f_country) in applicable
            ]}
            for r in results
        ]
        results = [r for r in results if r['activities']]

    countries = structure.get('countries', [])
    display_countries = [c for c in countries if not f_country or c['code'] == f_country]

    responsible_keys = set()
    if f_responsible and not error:
        _, submissions_raw, _, _ = _load(uid)
        fp = module.FIELD_PATHS
        for sub in submissions_raw:
            act_code = sub.get(fp['activity_code'], '')
            country = sub.get(fp['country'], '')
            responsible = sub.get(fp['activity_responsible'], '').strip()
            main = module.extract_main_activity(act_code)
            if responsible == f_responsible and main and country:
                responsible_keys.add((main, country))

    form_name = uid
    try:
        form_name = cache_helpers.get_cached(
            cache_helpers.schema_key(uid),
            lambda: api_client.get_schema(uid, _config()),
        ).get('name', uid)
    except Exception:
        pass

    matrix = []
    for result in results:
        rows = []
        for activity in result['activities']:
            cells = []
            for c in display_countries:
                is_applicable = (activity['code'], c['code']) in applicable
                count = coverage_data.get((activity['code'], c['code']), 0)
                responsible_match = (
                    not f_responsible
                    or (activity['code'], c['code']) in responsible_keys
                )
                sub_url = (
                    f'/dashboard/{uid}/submissions/'
                    f'?activity={activity["code"]}&country={c["code"]}'
                    + (f'&responsible={f_responsible}' if f_responsible else '')
                )
                cells.append({
                    'country': c['code'],
                    'applicable': is_applicable,
                    'count': count,
                    'responsible_match': responsible_match,
                    'sub_url': sub_url,
                })
            rows.append({'code': activity['code'], 'label': activity['label'], 'cells': cells})
        matrix.append({'result': result, 'rows': rows})

    return render(request, 'dashboard/coverage.html', {
        'uid': uid,
        'form_name': form_name,
        'form_label': module.form_label if module else '',
        'matrix': matrix,
        'all_results': structure.get('results', []),
        'countries': countries,
        'display_countries': display_countries,
        'responsibles': responsibles,
        'f_country': f_country,
        'f_result': f_result,
        'f_activity': f_activity,
        'f_responsible': f_responsible,
        'error': error,
    })


# ── JSON dashboard viewer ──────────────────────────────────────────────────────

@login_required
def view_dashboard(request, uid, pk):
    if not _user_can_access_form(request.user, uid):
        return redirect('/dashboard/')
    form = get_object_or_404(ConfiguredForm, uid=uid)
    dash_config = get_object_or_404(DashboardConfig, pk=pk, form=form)
    return _render_generic_dashboard(request, uid, form, dash_config)


# ── Dashboard editor ───────────────────────────────────────────────────────────

def _build_widget_from_post(post):
    wtype = post.get('type', 'summary_stat')
    widget = {'type': wtype, 'title': post.get('title', '').strip()}
    if wtype == 'pie_chart':
        widget['field'] = post.get('field', '').strip()
        pie_vals   = post.getlist('pie_val')
        pie_colors = post.getlist('pie_color')
        pie_labels = post.getlist('pie_label')
        pie_meta = {}
        for i, val in enumerate(pie_vals):
            if not val:
                continue
            entry = {}
            if i < len(pie_colors) and pie_colors[i]:
                entry['color'] = pie_colors[i]
            if i < len(pie_labels) and pie_labels[i].strip():
                entry['label'] = pie_labels[i].strip()
            if entry:
                pie_meta[val] = entry
        if pie_meta:
            widget['pie_meta'] = pie_meta
    elif wtype == 'summary_stat':
        widget['field'] = post.get('field', '').strip() or None
        widget['aggregation'] = post.get('aggregation', 'count')
    elif wtype == 'data_table':
        if post.get('pivot_mode') == '1':
            widget['pivot_mode'] = True
            widget['pivot_row_field'] = post.get('pivot_row_field', '').strip()
            widget['pivot_col_field'] = post.get('pivot_col_field', '').strip()
        else:
            widget['fields'] = [f.strip() for f in post.getlist('fields') if f.strip()]
    elif wtype == 'grouped_chart':
        widget['field'] = post.get('field', '').strip()
        widget['chart_style'] = post.get('chart_style', 'bar')
        widget['stacked'] = post.get('stacked') == '1'
        # Build series as ordered list of {field, color, label}
        paths  = post.getlist('series_path')
        colors = post.getlist('series_color')
        labels = post.getlist('series_label')
        series = []
        for p, c, l in zip(paths, colors, labels):
            if p:
                entry = {'field': p}
                if c:
                    entry['color'] = c
                if l.strip():
                    entry['label'] = l.strip()
                series.append(entry)
        widget['series'] = series
        # Per-value colors for select_one series fields
        svc_fields = post.getlist('svc_field')
        svc_vals   = post.getlist('svc_val')
        svc_colors = post.getlist('svc_color')
        series_value_colors = {}
        for f, v, c in zip(svc_fields, svc_vals, svc_colors):
            if f and v:
                series_value_colors.setdefault(f, {})[v] = c
        if series_value_colors:
            widget['series_value_colors'] = series_value_colors

    # Per-widget filters (all widget types)
    wf_fields = post.getlist('wf_field')
    wf_ops    = post.getlist('wf_op')
    wf_vals   = post.getlist('wf_val')
    wf_labels = post.getlist('wf_label')
    widget_filters = []
    for f, op, v, lbl in zip(wf_fields, wf_ops, wf_vals, wf_labels):
        if f:
            widget_filters.append({'field': f, 'op': op or 'eq', 'value': v, 'label': lbl or f})
    if widget_filters:
        widget['widget_filters'] = widget_filters

    return widget


@login_required
def dashboard_editor_list(request, uid):
    """List all JSON dashboards for a form and create new ones."""
    can_edit = _is_power_user(request.user) or \
        uid in _user_admin_forms(request.user).values_list('uid', flat=True)
    if not can_edit:
        return redirect('/dashboard/')

    form = get_object_or_404(ConfiguredForm, uid=uid)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            dc = DashboardConfig.objects.create(
                form=form,
                name=name,
                schema_version=1,
                config={'schema_version': 1, 'rows': []},
            )
            return redirect(f'/dashboard/{uid}/editor/{dc.pk}/')

    configs = DashboardConfig.objects.filter(form=form)
    return render(request, 'dashboard/editor_list.html', {
        'uid': uid,
        'form': form,
        'configs': configs,
    })


@login_required
def dashboard_editor(request, uid, pk):
    can_edit = _is_power_user(request.user) or \
        uid in _user_admin_forms(request.user).values_list('uid', flat=True)
    if not can_edit:
        return redirect('/dashboard/')

    form = get_object_or_404(ConfiguredForm, uid=uid)
    dash_config = get_object_or_404(DashboardConfig, pk=pk, form=form)
    config_json = copy.deepcopy(dash_config.config or {'schema_version': 1, 'rows': []})
    if 'rows' not in config_json:
        config_json['rows'] = []

    if request.method == 'POST':
        action = request.POST.get('action', '')
        rows = config_json.get('rows', [])

        if action == 'add_row':
            rows.append({'columns': 1, 'widgets': []})

        elif action == 'delete_row':
            idx = _safe_int(request.POST.get('row_idx'))
            if idx is not None and 0 <= idx < len(rows):
                rows.pop(idx)

        elif action == 'move_row_up':
            idx = _safe_int(request.POST.get('row_idx'))
            if idx is not None and idx > 0:
                rows[idx - 1], rows[idx] = rows[idx], rows[idx - 1]

        elif action == 'move_row_down':
            idx = _safe_int(request.POST.get('row_idx'))
            if idx is not None and idx < len(rows) - 1:
                rows[idx + 1], rows[idx] = rows[idx], rows[idx + 1]

        elif action == 'set_columns':
            idx = _safe_int(request.POST.get('row_idx'))
            cols = _safe_int(request.POST.get('columns'))
            if idx is not None and cols in (1, 2, 3) and 0 <= idx < len(rows):
                rows[idx]['columns'] = cols

        elif action == 'add_widget':
            idx = _safe_int(request.POST.get('row_idx'))
            if idx is not None and 0 <= idx < len(rows):
                rows[idx]['widgets'].append(_build_widget_from_post(request.POST))
                # Auto-adjust columns to match widget count (up to 3)
                rows[idx]['columns'] = min(len(rows[idx]['widgets']), 3)

        elif action == 'edit_widget':
            ridx = _safe_int(request.POST.get('row_idx'))
            widx = _safe_int(request.POST.get('widget_idx'))
            if ridx is not None and widx is not None and \
                    0 <= ridx < len(rows) and 0 <= widx < len(rows[ridx]['widgets']):
                rows[ridx]['widgets'][widx] = _build_widget_from_post(request.POST)

        elif action == 'delete_widget':
            ridx = _safe_int(request.POST.get('row_idx'))
            widx = _safe_int(request.POST.get('widget_idx'))
            if ridx is not None and widx is not None and \
                    0 <= ridx < len(rows) and 0 <= widx < len(rows[ridx]['widgets']):
                rows[ridx]['widgets'].pop(widx)

        elif action == 'rename':
            new_name = request.POST.get('name', '').strip()
            if new_name:
                dash_config.name = new_name
                dash_config.save()
            return redirect(f'/dashboard/{uid}/editor/{pk}/')

        elif action == 'delete_config':
            dash_config.delete()
            return redirect(f'/dashboard/{uid}/editor/')

        elif action == 'add_filter':
            field = request.POST.get('filter_field', '').strip()
            label = request.POST.get('filter_label', '').strip()
            if field:
                filters = config_json.setdefault('filters', [])
                if not any(f.get('field') == field for f in filters):
                    filters.append({'field': field, 'label': label or field})

        elif action == 'delete_filter':
            fidx = _safe_int(request.POST.get('filter_idx'))
            filters = config_json.get('filters', [])
            if fidx is not None and 0 <= fidx < len(filters):
                filters.pop(fidx)
            config_json['filters'] = filters

        elif action == 'move_filter_up':
            fidx = _safe_int(request.POST.get('filter_idx'))
            filters = config_json.get('filters', [])
            if fidx is not None and fidx > 0:
                filters[fidx - 1], filters[fidx] = filters[fidx], filters[fidx - 1]
            config_json['filters'] = filters

        elif action == 'move_filter_down':
            fidx = _safe_int(request.POST.get('filter_idx'))
            filters = config_json.get('filters', [])
            if fidx is not None and fidx < len(filters) - 1:
                filters[fidx + 1], filters[fidx] = filters[fidx], filters[fidx + 1]
            config_json['filters'] = filters

        config_json['rows'] = rows
        dash_config.config = config_json
        dash_config.save()
        return redirect(f'/dashboard/{uid}/editor/{pk}/')

    # Annotate rows/widgets with indices for template use
    rows_ctx = []
    for i, row in enumerate(config_json.get('rows', [])):
        widgets_ctx = []
        for j, widget in enumerate(row.get('widgets', [])):
            w = dict(widget)
            # Backward compat: series list of strings → list of dicts
            raw_series = w.get('series') or ([w['field2']] if w.get('field2') else [])
            series_meta = {}
            series_paths = []
            for idx, entry in enumerate(raw_series):
                if isinstance(entry, dict):
                    path = entry.get('field', '')
                    series_meta[path] = {
                        'color': entry.get('color') or _WIDGET_COLORS[idx % len(_WIDGET_COLORS)],
                        'label': entry.get('label', ''),
                    }
                else:
                    path = entry
                    series_meta[path] = {
                        'color': _WIDGET_COLORS[idx % len(_WIDGET_COLORS)],
                        'label': '',
                    }
                if path:
                    series_paths.append(path)
            w['series'] = raw_series
            w['series_meta'] = series_meta
            w['series_paths'] = series_paths
            w['pie_meta'] = w.get('pie_meta', {})
            w['widget_filters'] = w.get('widget_filters', [])
            widgets_ctx.append({**w, 'widget_idx': j, 'edit_key': f'{i}-{j}'})
        cols = max(row.get('columns', 1), min(len(widgets_ctx), 3))
        rows_ctx.append({**row, 'row_idx': i, 'columns': cols, 'widgets': widgets_ctx})

    field_choices = []
    pie_choices = {}
    has_schema = False
    try:
        kobo_config = _config()
        schema = cache_helpers.get_cached(
            cache_helpers.schema_key(uid),
            lambda: api_client.get_schema(uid, kobo_config),
            ttl=form.cache_ttl_seconds,
        )
        field_choices = api_client.get_field_choices(schema)
        pie_choices = {
            path: list(api_client.get_choice_labels(schema, path).items())
            for path, lbl, ftype in field_choices
            if ftype.startswith('select_one') and api_client.get_choice_labels(schema, path)
        }
        has_schema = True
    except api_client.KoboAPIError:
        pass

    # Second pass: enrich widget dicts for the editor UI
    field_label_map = {p: lbl for p, lbl, ft in field_choices}
    for row in rows_ctx:
        for w in row['widgets']:
            # Ordered display list for the series editor
            w['series_display'] = [
                {
                    'path': path,
                    'field_label': field_label_map.get(path, path),
                    'color': w['series_meta'].get(path, {}).get('color', ''),
                    'label': w['series_meta'].get(path, {}).get('label', ''),
                }
                for path in w.get('series_paths', [])
            ]
            # Per-value color pickers for select_one series fields
            svc_store = w.get('series_value_colors') or {}
            svc_list = []
            for path in w.get('series_paths', []):
                if path in pie_choices:
                    path_colors = svc_store.get(path, {}) if isinstance(svc_store, dict) else {}
                    svc_list.append({
                        'field': path,
                        'field_label': field_label_map.get(path, path),
                        'choices': [
                            {'val': val, 'label': val_label,
                             'color': path_colors.get(val, '')}
                            for val, val_label in pie_choices[path]
                        ],
                    })
            w['series_value_choices'] = svc_list

    filters_ctx = [
        {**f, 'filter_idx': i}
        for i, f in enumerate(config_json.get('filters', []))
    ]

    return render(request, 'dashboard/editor.html', {
        'uid': uid,
        'form': form,
        'dash_config': dash_config,
        'rows': rows_ctx,
        'filters': filters_ctx,
        'field_choices': field_choices,
        'pie_choices': pie_choices,
        'has_schema': has_schema,
        'empty_widget': {},
    })


def _safe_int(val):
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


# ── Submission list ─────────────────────────────────────────────────────────────

@login_required
def submission_list(request, uid):
    if not _user_can_access_form(request.user, uid):
        return redirect('/dashboard/')
    error = None
    parsed_submissions = []
    structure = {}

    f_country = request.GET.get('country', '')
    f_activity = request.GET.get('activity', '')
    f_responsible = request.GET.get('responsible', '')

    try:
        schema, submissions, structure, module = _load(uid)
        if module is None or hasattr(module, 'parse_submissions') \
                or 'activity_code' not in module.FIELD_PATHS:
            return form_detail(request, uid)

        fp = module.FIELD_PATHS
        for sub in submissions:
            act_code = sub.get(fp.get('activity_code', ''), '')
            country = sub.get(fp.get('country', ''), '')
            main = module.extract_main_activity(act_code)
            responsible = sub.get(fp.get('activity_responsible', ''), '').strip()

            if f_activity and main != f_activity:
                continue
            if f_country and country != f_country:
                continue
            if f_responsible and responsible != f_responsible:
                continue

            parsed_submissions.append(module.parse_submission_detail(sub, structure))

    except api_client.KoboAPIError as exc:
        error = str(exc)

    paginator = Paginator(parsed_submissions, PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get('page'))

    activity_label = ''
    if f_activity:
        activity_label = structure.get('activity_labels', {}).get(f_activity, f_activity)

    return render(request, 'dashboard/submission_list.html', {
        'uid': uid,
        'page_obj': page_obj,
        'total': len(parsed_submissions),
        'f_country': f_country,
        'f_activity': f_activity,
        'f_responsible': f_responsible,
        'activity_label': activity_label,
        'country_label': structure.get('country_labels', {}).get(f_country, f_country),
        'error': error,
    })


# ── Submission detail ───────────────────────────────────────────────────────────

@login_required
def submission_detail(request, uid, sub_id):
    if not _user_can_access_form(request.user, uid):
        return redirect('/dashboard/')
    error = None
    parsed = None

    try:
        schema, submissions, structure, module = _load(uid)
        if module is None or hasattr(module, 'parse_submissions') \
                or 'activity_code' not in module.FIELD_PATHS:
            return form_detail(request, uid)

        raw = next((s for s in submissions if s.get('_id') == sub_id), None)
        if raw is None:
            error = f'Submission #{sub_id} not found.'
        else:
            parsed = module.parse_submission_detail(raw, structure)
    except api_client.KoboAPIError as exc:
        error = str(exc)

    back_url = request.GET.get('back', f'/dashboard/{uid}/')

    return render(request, 'dashboard/submission_detail.html', {
        'uid': uid,
        'parsed': parsed,
        'back_url': back_url,
        'error': error,
    })


# ── Help ───────────────────────────────────────────────────────────────────────

@login_required
def manual(request):
    lang = getattr(request, 'LANGUAGE_CODE', 'fr')[:2]
    templates = {
        'en': 'dashboard/manual_en.html',
        'es': 'dashboard/manual_es.html',
        'ar': 'dashboard/manual_ar.html',
        'ru': 'dashboard/manual_ru.html',
    }
    return render(request, templates.get(lang, 'dashboard/manual.html'), {})


# ── Refresh ────────────────────────────────────────────────────────────────────

@login_required
def refresh_form(request, uid):
    if not _user_can_access_form(request.user, uid):
        return redirect('/dashboard/')
    cache_helpers.invalidate(cache_helpers.schema_key(uid))
    cache_helpers.invalidate(cache_helpers.submissions_key(uid))
    cache_helpers.invalidate(f'kobo_structure_{uid}')
    return redirect(request.GET.get('next', f'/dashboard/{uid}/'))


# ── Exports ────────────────────────────────────────────────────────────────────

class _Echo:
    def write(self, value):
        return value


def _module_csv_rows(writer, submissions, module):
    yield writer.writerow(module.EXPORT_HEADERS)
    i = 1
    for sub in submissions:
        ps = module.parse_submissions([sub])[0]
        if not ps['indicators']:
            yield writer.writerow([
                i, ps['country_label'], ps['year'], ps['quarter'], ps['reporter'],
                '', '', '', 0,
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
            ])
            i += 1
        else:
            for ind in ps['indicators']:
                age = ind['age']
                dis = ind['disability']
                sta = ind['status']
                yield writer.writerow([
                    i, ps['country_label'], ps['year'], ps['quarter'], ps['reporter'],
                    ind['result_label'], ind['code'], ind['label'], ind['total'],
                    age.get('male_total', ''), age.get('fem_total', ''),
                    age.get('male_0_5', ''), age.get('male_6_18', ''),
                    age.get('male_19_49', ''), age.get('male_50p', ''),
                    age.get('fem_0_5', ''), age.get('fem_6_18', ''),
                    age.get('fem_19_49', ''), age.get('fem_50p', ''),
                    dis.get('with', ''), dis.get('without', ''),
                    sta.get('pdi', ''), sta.get('host', ''), sta.get('refugee', ''),
                    sta.get('returnees', ''), sta.get('stateless', ''), sta.get('other', ''),
                ])
                i += 1


@login_required
def export_csv(request, uid):
    if not _user_can_access_form(request.user, uid):
        return redirect('/dashboard/')
    try:
        schema, submissions, structure, module = _load(uid)
    except api_client.KoboAPIError as exc:
        return HttpResponse(f'API error: {exc}', status=502)

    if module is None:
        return HttpResponse('Export non disponible sans module de formulaire.', status=400)

    pseudo_buffer = _Echo()
    writer = csv.writer(pseudo_buffer)

    if hasattr(module, 'parse_submissions'):
        response = StreamingHttpResponse(
            _module_csv_rows(writer, submissions, module),
            content_type='text/csv; charset=utf-8-sig',
        )
        response['Content-Disposition'] = f'attachment; filename="{uid}_donnees.csv"'
        return response

    def rows():
        yield writer.writerow(module.EXPORT_HEADERS)
        i = 1
        for sub in submissions:
            p = module.parse_submission_detail(sub, structure)
            act = p['activity']
            if not p['risks']:
                yield writer.writerow([
                    i, act['country_label'], act['activity_location'],
                    act['activity_code'], act['activity_label'],
                    act['activity_responsible'], act['start_date'], act['end_date'],
                    '', '', '',
                ])
                i += 1
            else:
                for risk in p['risks']:
                    yield writer.writerow([
                        i, act['country_label'], act['activity_location'],
                        act['activity_code'], act['activity_label'],
                        act['activity_responsible'], act['start_date'], act['end_date'],
                        risk['category_label'], risk['description'],
                        ' | '.join(risk['measures']),
                    ])
                    i += 1

    response = StreamingHttpResponse(rows(), content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{uid}_donnees.csv"'
    return response


def _module_xlsx(submissions, module):
    """Build openpyxl Workbook for indicator-module export."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Données'

    header_fill = PatternFill('solid', fgColor='C00000')
    ws.append(module.EXPORT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)

    i = 1
    for sub in submissions:
        ps = module.parse_submissions([sub])[0]
        if not ps['indicators']:
            ws.append([i, ps['country_label'], ps['year'], ps['quarter'], ps['reporter'],
                       '', '', '', 0, *([''] * 18)])
            i += 1
        else:
            for ind in ps['indicators']:
                age = ind['age']
                dis = ind['disability']
                sta = ind['status']
                ws.append([
                    i, ps['country_label'], ps['year'], ps['quarter'], ps['reporter'],
                    ind['result_label'], ind['code'], ind['label'], ind['total'],
                    age.get('male_total', ''), age.get('fem_total', ''),
                    age.get('male_0_5', ''), age.get('male_6_18', ''),
                    age.get('male_19_49', ''), age.get('male_50p', ''),
                    age.get('fem_0_5', ''), age.get('fem_6_18', ''),
                    age.get('fem_19_49', ''), age.get('fem_50p', ''),
                    dis.get('with', ''), dis.get('without', ''),
                    sta.get('pdi', ''), sta.get('host', ''), sta.get('refugee', ''),
                    sta.get('returnees', ''), sta.get('stateless', ''), sta.get('other', ''),
                ])
                i += 1

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    return wb


@login_required
def export_xlsx(request, uid):
    if not _user_can_access_form(request.user, uid):
        return redirect('/dashboard/')
    try:
        schema, submissions, structure, module = _load(uid)
    except api_client.KoboAPIError as exc:
        return HttpResponse(f'API error: {exc}', status=502)

    if module is None:
        return HttpResponse('Export non disponible sans module de formulaire.', status=400)

    if hasattr(module, 'parse_submissions'):
        wb = _module_xlsx(submissions, module)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{uid}_donnees.xlsx"'
        return response

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Données'

    ws.append(module.EXPORT_HEADERS)
    header_fill = PatternFill('solid', fgColor='C00000')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)

    i = 1
    for sub in submissions:
        p = module.parse_submission_detail(sub, structure)
        act = p['activity']
        if not p['risks']:
            ws.append([
                i, act['country_label'], act['activity_location'],
                act['activity_code'], act['activity_label'],
                act['activity_responsible'], act['start_date'], act['end_date'],
                '', '', '',
            ])
            i += 1
        else:
            for risk in p['risks']:
                ws.append([
                    i, act['country_label'], act['activity_location'],
                    act['activity_code'], act['activity_label'],
                    act['activity_responsible'], act['start_date'], act['end_date'],
                    risk['category_label'], risk['description'],
                    '\n'.join(risk['measures']),
                ])
                i += 1

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{uid}_donnees.xlsx"'
    return response


# ── User management (staff only) ───────────────────────────────────────────────

def _staff_required(view_fn):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_staff:
            return redirect('/dashboard/')
        return view_fn(request, *args, **kwargs)
    wrapper.__name__ = view_fn.__name__
    return wrapper


@_staff_required
def user_list(request):
    User = get_user_model()
    pending = User.objects.filter(is_active=False).select_related('profile').order_by('date_joined')
    active = User.objects.filter(is_active=True, is_superuser=False).select_related('profile').order_by('date_joined')
    return render(request, 'dashboard/user_list.html', {
        'pending': pending,
        'active': active,
        'power_user_emails': django_settings.POWER_USER_EMAILS,
    })


@_staff_required
def user_activate(request, user_id):
    if request.method == 'POST':
        User = get_user_model()
        user = get_object_or_404(User, pk=user_id)
        user.is_active = True
        user.save()
    return redirect('/dashboard/users/')


@_staff_required
def user_deactivate(request, user_id):
    if request.method == 'POST':
        User = get_user_model()
        user = get_object_or_404(User, pk=user_id)
        if user != request.user:
            user.is_active = False
            user.save()
    return redirect('/dashboard/users/')


@_staff_required
def user_delete(request, user_id):
    if request.method == 'POST':
        User = get_user_model()
        user = get_object_or_404(User, pk=user_id)
        if user != request.user:
            user.delete()
    return redirect('/dashboard/users/')


@login_required
def user_toggle_power(request, user_id):
    if request.method != 'POST' or not _is_power_user(request.user):
        return redirect('/dashboard/')
    User = get_user_model()
    user = get_object_or_404(User, pk=user_id)
    if user != request.user and user.email not in django_settings.POWER_USER_EMAILS:
        profile = user.profile
        profile.is_power_user = not profile.is_power_user
        profile.save()
    return redirect('/dashboard/users/')


@login_required
def generate_reset_link(request, user_id):
    if request.method != 'POST':
        return redirect('/dashboard/')
    if not _can_manage_user(request.user, user_id):
        return redirect('/dashboard/')
    User = get_user_model()
    user = get_object_or_404(User, pk=user_id)
    uid = urlsafe_base64_encode(force_bytes(user.pk))
    token = default_token_generator.make_token(user)
    link = request.build_absolute_uri(f'/accounts/password-reset/confirm/{uid}/{token}/')
    back = '/dashboard/my-group/' if _is_group_admin(request.user) and not _is_power_user(request.user) else '/dashboard/users/'
    return render(request, 'dashboard/password_reset_link.html', {
        'target_user': user,
        'link': link,
        'back_url': back,
    })


# ── Group management (power user only) ────────────────────────────────────────

@login_required
def group_edit(request, group_id):
    if not _is_power_user(request.user):
        return redirect('/dashboard/')
    group = get_object_or_404(DashboardGroup, pk=group_id)
    User = get_user_model()
    all_forms = ConfiguredForm.objects.all()
    all_users = User.objects.filter(is_active=True, is_superuser=False).select_related('profile')

    if request.method == 'POST':
        group.name = request.POST.get('group_name', group.name).strip() or group.name
        group.save()
        group.forms.set(all_forms.filter(pk__in=request.POST.getlist('form_ids')))
        group.members.set(all_users.filter(pk__in=request.POST.getlist('member_ids')))
        group.admins.set(all_users.filter(pk__in=request.POST.getlist('admin_ids')))
        return redirect('/dashboard/settings/')

    return render(request, 'dashboard/group_edit.html', {
        'group': group,
        'all_forms': all_forms,
        'all_users': all_users,
    })


# ── Group admin view ──────────────────────────────────────────────────────────

@login_required
def my_group(request, user_id=None):
    if not _is_group_admin(request.user) and not _is_power_user(request.user):
        return redirect('/dashboard/')

    User = get_user_model()

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'add_member':
            group_id = request.POST.get('group_id', '').strip()
            new_user_id = request.POST.get('new_user_id', '').strip()
            group = DashboardGroup.objects.filter(pk=group_id, admins=request.user).first()
            if group and new_user_id:
                group.members.add(new_user_id)

        elif user_id is not None and _can_manage_user(request.user, user_id):
            target = get_object_or_404(User, pk=user_id)
            if action == 'deactivate' and target != request.user:
                target.is_active = False
                target.save()
            elif action == 'delete' and target != request.user:
                target.delete()

        return redirect('/dashboard/my-group/')

    groups = DashboardGroup.objects.filter(admins=request.user).prefetch_related(
        'members__profile', 'forms'
    )
    # Active users not already in any of the admin's groups (per group computed in template)
    all_active_users = User.objects.filter(is_active=True, is_superuser=False).select_related('profile')
    return render(request, 'dashboard/my_group.html', {
        'groups': groups,
        'all_active_users': all_active_users,
    })
